import logging
import re
import asyncio
import signal
import io
from datetime import datetime, date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, Document
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)

# ===================== SETTINGS =====================
BOT_TOKEN = "8302225372:AAE1aLrBi15j066O5eAWPyIb-PkRa_Zw_vQ"  # YANGI TOKEN QO'YING
ADMIN_ID = 8013467870

QUESTION_TIMEOUT_SEC = 120   # har savolga 2 daqiqa
TICK_SEC = 10                # har 10 sekundda timer update

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
EXCEL_PATH = DATA_DIR / "applications.xlsx"

# ===================== LOGGING =====================
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ===================== STATES =====================
(SAVOL1, SAVOL2, SAVOL3, SAVOL4, SAVOL5, SAVOL6,
 ISM_FAMILIYA, YOSH, SHAHAR, TEL, OXIRGI_ISH, SOHALAR, TAJRIBA_YIL, VAZIFALAR, BOSHlash, MAOSH) = range(16)

# ===================== GLOBALS =====================
user_scores = {}  # closer score (0..18)

REGIONS = [
    "Toshkent",
    "Samarqand",
    "Farg'ona",
    "Andijon",
    "Namangan",
    "Buxoro",
    "Xorazm",
    "Qashqadaryo",
    "Surxondaryo",
    "Jizzax",
    "Sirdaryo",
    "Navoiy",
    "Qoraqalpog'iston",
    "Boshqa",
]

ABOUT_TEXT = (
    "🔥 BUYUK ZAMON — Sotuv bo'limi\n\n"
    "💼 High-Ticket Closer testi\n"
    "⏰ 10:00–18:00 | 🍽 13:00–14:00 (o'z hisobidan)\n"
    "📆 Haftada 6 kun | 🎯 19–38 yosh\n"
    "💰 Har sotuvdan 5% daromad"
)

# ===================== EXCEL =====================
EXCEL_HEADERS = [
    "timestamp",
    "status",
    "tg_user_id",
    "tg_username",
    "ism_familiya",
    "yosh",
    "shahar",
    "tel",
    "oxirgi_ish",
    "sohalar",
    "tajriba_yil",
    "vazifalar",
    "boshlash",
    "maosh",
    "interview_date",
    "closer_score",
]

def ensure_excel_file():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if EXCEL_PATH.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "DATA"
    ws.append(EXCEL_HEADERS)
    # Header styling
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    wb.save(EXCEL_PATH)

def excel_append_row(row: list):
    try:
        ensure_excel_file()
        wb = load_workbook(EXCEL_PATH)
        ws = wb["DATA"] if "DATA" in wb.sheetnames else wb.active
        ws.append(row)
        wb.save(EXCEL_PATH)
        return True
    except Exception as e:
        logger.exception("Excel append error: %s", e)
        return False

async def send_excel_to_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin ga Excel faylni yuborish"""
    user_id = update.effective_user.id
    if user_id != ADMIN_ID:
        await update.message.reply_text("⛔️ Siz admin emassiz!")
        return
    
    try:
        ensure_excel_file()
        with open(EXCEL_PATH, "rb") as f:
            await context.bot.send_document(
                chat_id=ADMIN_ID,
                document=f,
                filename=f"arizalar_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx",
                caption=f"📊 Arizalar bazasi\nJami: {get_applications_count()} ta ariza"
            )
    except Exception as e:
        logger.exception("Excel send error: %s", e)
        await update.message.reply_text(f"❌ Xato: {str(e)}")

def get_applications_count():
    try:
        ensure_excel_file()
        wb = load_workbook(EXCEL_PATH)
        ws = wb["DATA"] if "DATA" in wb.sheetnames else wb.active
        return ws.max_row - 1
    except:
        return 0

# ===================== LOCK / EXPIRE =====================
def _locks(context: ContextTypes.DEFAULT_TYPE) -> dict:
    return context.application.bot_data.setdefault("locks", {})

def _expired(context: ContextTypes.DEFAULT_TYPE) -> dict:
    return context.application.bot_data.setdefault("expired", {})

def is_locked(context: ContextTypes.DEFAULT_TYPE, uid: int) -> bool:
    lock_until = _locks(context).get(uid)
    return bool(lock_until and date.today() < lock_until)

def lock_until_tomorrow(context: ContextTypes.DEFAULT_TYPE, uid: int):
    _locks(context)[uid] = date.today() + timedelta(days=1)

def set_expired(context: ContextTypes.DEFAULT_TYPE, uid: int, val: bool):
    _expired(context)[uid] = val

def is_expired(context: ContextTypes.DEFAULT_TYPE, uid: int) -> bool:
    return bool(_expired(context).get(uid, False))

def guard_expired(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    uid = update.effective_user.id if update.effective_user else None
    if not uid:
        return True
    return is_expired(context, uid)

# ===================== TIMER =====================
def cancel_timer(user_data: dict):
    task = user_data.get("timer_task")
    if task and not task.done():
        task.cancel()
    user_data["timer_task"] = None

async def _safe_edit(context: ContextTypes.DEFAULT_TYPE, chat_id: int, message_id: int, text: str, reply_markup):
    try:
        await context.bot.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=text,
            reply_markup=reply_markup,
        )
    except Exception:
        pass

async def timer_task_fn(
    context: ContextTypes.DEFAULT_TYPE,
    uid: int,
    chat_id: int,
    message_id: int,
    base_text: str,
    reply_markup,
):
    left = QUESTION_TIMEOUT_SEC
    try:
        while left > 0:
            await _safe_edit(
                context,
                chat_id,
                message_id,
                base_text + f"\n\n⏳ Qoldi: {left}s",
                reply_markup,
            )
            await asyncio.sleep(TICK_SEC)
            left -= TICK_SEC

        if is_expired(context, uid):
            return

        set_expired(context, uid, True)
        lock_until_tomorrow(context, uid)

        try:
            await context.bot.edit_message_reply_markup(chat_id=chat_id, message_id=message_id, reply_markup=None)
        except Exception:
            pass

        ud = context.application.user_data.get(uid, {}) if hasattr(context.application, "user_data") else {}
        closer_score = user_scores.get(uid, 0)

        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "TIMEOUT",
            str(uid),
            ud.get("username", ""),
            ud.get("ism_familiya", ""),
            ud.get("yosh", ""),
            ud.get("shahar", ""),
            ud.get("tel", ""),
            ud.get("oxirgi_ish", ""),
            ud.get("sohalar", ""),
            ud.get("tajriba_yil", ""),
            ud.get("vazifalar", ""),
            ud.get("boshlash", ""),
            ud.get("maosh", ""),
            ud.get("interview_date", ""),
            f"{closer_score}/18",
        ]
        excel_append_row(row)

        await context.bot.send_message(
            chat_id=chat_id,
            text="⏰ Vaqt tugadi! Test bekor qilindi.\n\nErtaga yana urinib ko'ring. /start"
        )

    except asyncio.CancelledError:
        return

def start_timer(context: ContextTypes.DEFAULT_TYPE, user_data: dict, uid: int, chat_id: int, message_id: int, base_text: str, reply_markup):
    cancel_timer(user_data)
    user_data["timer_task"] = asyncio.create_task(
        timer_task_fn(context, uid, chat_id, message_id, base_text, reply_markup)
    )

# ===================== HELPERS =====================
def build_regions_keyboard():
    rows, row = [], []
    for r in REGIONS:
        row.append(InlineKeyboardButton(r, callback_data=f"reg:{r}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    return InlineKeyboardMarkup(rows)

def normalize_uz_phone(text: str) -> str | None:
    digits = re.sub(r"\D", "", text.strip())
    if digits.startswith("998") and len(digits) == 12:
        return "+" + digits
    if len(digits) == 9:
        return "+998" + digits
    return None

def interview_date_keyboard():
    d1 = date.today() + timedelta(days=1)
    d2 = date.today() + timedelta(days=2)
    d3 = date.today() + timedelta(days=3)
    rows = [
        [InlineKeyboardButton(d1.strftime("%d.%m.%Y"), callback_data=f"int:{d1.isoformat()}")],
        [InlineKeyboardButton(d2.strftime("%d.%m.%Y"), callback_data=f"int:{d2.isoformat()}")],
        [InlineKeyboardButton(d3.strftime("%d.%m.%Y"), callback_data=f"int:{d3.isoformat()}")],
    ]
    return InlineKeyboardMarkup(rows)

# ===================== START =====================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [[InlineKeyboardButton("🚀 TESTNI BOSHLASH", callback_data="start_test")]]
    await update.message.reply_text(
        ABOUT_TEXT + "\n\n🚀 Testni boshlash uchun tugmani bosing:",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return ConversationHandler.END

# ===================== CLOSER TEST (6 SAVOL) =====================
async def start_test(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id

    if is_locked(context, uid):
        await query.edit_message_text("⛔️ Bugun urinish tugagan.\nErtaga yana urinib ko'ring. /start")
        return ConversationHandler.END

    context.user_data.clear()
    context.user_data["username"] = (query.from_user.username or "")
    user_scores[uid] = 0
    cancel_timer(context.user_data)
    set_expired(context, uid, False)

    keyboard = [
        [InlineKeyboardButton("Ha, katta cheklar yopganman", callback_data="s1_a")],
        [InlineKeyboardButton("Kichik cheklar bilan", callback_data="s1_b")],
        [InlineKeyboardButton("O'rganishga tayyorman", callback_data="s1_c")],
    ]
    rm = InlineKeyboardMarkup(keyboard)
    base_text = "❓ SAVOL 1/6\n\nSotuv qila olasizmi?"
    await query.edit_message_text(base_text + f"\n\n⏳ Qoldi: {QUESTION_TIMEOUT_SEC}s", reply_markup=rm)
    start_timer(context, context.user_data, uid, query.message.chat_id, query.message.message_id, base_text, rm)
    return SAVOL1

async def savol1_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    cancel_timer(context.user_data)

    if query.data == "s1_a":
        user_scores[uid] += 3
    elif query.data == "s1_b":
        user_scores[uid] += 2
    elif query.data == "s1_c":
        user_scores[uid] += 1

    keyboard = [
        [InlineKeyboardButton("Nima to'xtatmoqda?", callback_data="s2_b")],
        [InlineKeyboardButton("Yana urinaman", callback_data="s2_d")],
        [InlineKeyboardButton("Qachon qo'ng'iroq qilay?", callback_data="s2_a")],
        [InlineKeyboardButton("Chegirma qilaman", callback_data="s2_c")],
    ]
    rm = InlineKeyboardMarkup(keyboard)
    base_text = "❓ SAVOL 2/6\n\nMijoz \"o'ylab ko'raman\" desa?"
    await query.edit_message_text(base_text + f"\n\n⏳ Qoldi: {QUESTION_TIMEOUT_SEC}s", reply_markup=rm)
    start_timer(context, context.user_data, uid, query.message.chat_id, query.message.message_id, base_text, rm)
    return SAVOL2

async def savol2_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    cancel_timer(context.user_data)

    if query.data == "s2_b":
        user_scores[uid] += 3
    elif query.data == "s2_d":
        user_scores[uid] += 2
    elif query.data == "s2_a":
        user_scores[uid] += 1

    keyboard = [
        [InlineKeyboardButton("10+ sotuv", callback_data="s3_a")],
        [InlineKeyboardButton("5–10 sotuv", callback_data="s3_b")],
        [InlineKeyboardButton("Hali yo'q", callback_data="s3_c")],
    ]
    rm = InlineKeyboardMarkup(keyboard)
    base_text = "❓ SAVOL 3/6\n\nOyiga nechta sotuv qilasiz?"
    await query.edit_message_text(base_text + f"\n\n⏳ Qoldi: {QUESTION_TIMEOUT_SEC}s", reply_markup=rm)
    start_timer(context, context.user_data, uid, query.message.chat_id, query.message.message_id, base_text, rm)
    return SAVOL3

async def savol3_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    cancel_timer(context.user_data)

    if query.data == "s3_a":
        user_scores[uid] += 3
    elif query.data == "s3_b":
        user_scores[uid] += 2

    keyboard = [
        [InlineKeyboardButton("Ha", callback_data="s4_a")],
        [InlineKeyboardButton("Yo'q", callback_data="s4_b")],
    ]
    rm = InlineKeyboardMarkup(keyboard)
    base_text = "❓ SAVOL 4/6\n\n10:00–18:00 ish vaqti mosmi?"
    await query.edit_message_text(base_text + f"\n\n⏳ Qoldi: {QUESTION_TIMEOUT_SEC}s", reply_markup=rm)
    start_timer(context, context.user_data, uid, query.message.chat_id, query.message.message_id, base_text, rm)
    return SAVOL4

async def savol4_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    cancel_timer(context.user_data)

    if query.data == "s4_a":
        user_scores[uid] += 3
    else:
        user_scores[uid] -= 5

    keyboard = [
        [InlineKeyboardButton("Ha", callback_data="s5_a")],
        [InlineKeyboardButton("Yo'q", callback_data="s5_b")],
    ]
    rm = InlineKeyboardMarkup(keyboard)
    base_text = "❓ SAVOL 5/6\n\n19–38 yosh oralig'idamisiz?"
    await query.edit_message_text(base_text + f"\n\n⏳ Qoldi: {QUESTION_TIMEOUT_SEC}s", reply_markup=rm)
    start_timer(context, context.user_data, uid, query.message.chat_id, query.message.message_id, base_text, rm)
    return SAVOL5

async def savol5_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    query = update.callback_query
    await query.answer()
    cancel_timer(context.user_data)

    if query.data == "s5_b":
        await query.edit_message_text("❌ Yosh mos emas.")
        return ConversationHandler.END

    rm = interview_date_keyboard()
    base_text = "❓ SAVOL 6/6\n\nSuhbat sanasini tanlang:"
    await query.edit_message_text(base_text + f"\n\n⏳ Qoldi: {QUESTION_TIMEOUT_SEC}s", reply_markup=rm)
    start_timer(context, context.user_data, query.from_user.id, query.message.chat_id, query.message.message_id, base_text, rm)
    return SAVOL6

async def savol6_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END

    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    cancel_timer(context.user_data)

    if not query.data.startswith("int:"):
        return SAVOL6

    iso = query.data.split("int:", 1)[1].strip()
    try:
        d = date.fromisoformat(iso)
        context.user_data["interview_date"] = d.strftime("%d.%m.%Y")
    except Exception:
        context.user_data["interview_date"] = ""

    user_scores[uid] += 3
    score = user_scores.get(uid, 0)

    if score >= 12:
        await query.edit_message_text(f"🎉 Tabriklaymiz! Ball: {score}/18\n\n✍️ Ismingiz va familiyangizni kiriting:")
        return ISM_FAMILIYA
    elif score >= 8:
        await query.edit_message_text(f"Ball: {score}/18\n\n✍️ Ismingiz va familiyangizni kiriting:")
        return ISM_FAMILIYA
    else:
        await query.edit_message_text(f"❌ Hozircha mos emas.\n\nBall: {score}/18")
        return ConversationHandler.END

# ===================== ANKETA SAVOLLARI =====================
async def get_ism_familiya(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    context.user_data["ism_familiya"] = update.message.text.strip()
    await update.message.reply_text("🎂 Necha yoshdasiz?")
    return YOSH

async def get_yosh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    yosh_txt = update.message.text.strip()
    if not yosh_txt.isdigit():
        await update.message.reply_text("❌ Faqat raqam kiriting. Masalan: 23")
        return YOSH
    context.user_data["yosh"] = yosh_txt
    await update.message.reply_text("📍 Qaysi shahardansiz?", reply_markup=build_regions_keyboard())
    return SHAHAR

async def get_shahar_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    query = update.callback_query
    await query.answer()
    cancel_timer(context.user_data)

    if not query.data.startswith("reg:"):
        return SHAHAR

    region = query.data.split("reg:", 1)[1].strip()
    context.user_data["shahar"] = region
    await query.edit_message_text("📞 Telefon raqamingizni kiriting:\nMisol: +998901234567 yoki 901234567")
    return TEL

async def get_tel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    phone = normalize_uz_phone(update.message.text)
    if not phone:
        await update.message.reply_text("❌ Telefon formati xato.\nMisol: +998901234567 yoki 901234567\nQayta yuboring:")
        return TEL
    context.user_data["tel"] = phone
    await update.message.reply_text("💼 Oxirgi ish joyingiz va lavozimingizni yozing:")
    return OXIRGI_ISH

async def get_oxirgi_ish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    context.user_data["oxirgi_ish"] = update.message.text.strip()
    await update.message.reply_text("🎯 Qaysi sohalarda ishlagansiz?")
    return SOHALAR

async def get_sohalar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    context.user_data["sohalar"] = update.message.text.strip()
    await update.message.reply_text("📊 Necha yillik ish tajribangiz bor?")
    return TAJRIBA_YIL

async def get_tajriba_yil(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    context.user_data["tajriba_yil"] = update.message.text.strip()
    await update.message.reply_text("📝 Oxirgi ish joyingizdagi asosiy vazifalaringizni yozing:")
    return VAZIFALAR

async def get_vazifalar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    context.user_data["vazifalar"] = update.message.text.strip()
    await update.message.reply_text("📅 Qachondan ish boshlay olasiz?")
    return BOSHlash

async def get_boshlash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    context.user_data["boshlash"] = update.message.text.strip()
    await update.message.reply_text("💰 Maosh bo'yicha kutganingizni yozing:")
    return MAOSH

async def get_maosh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if guard_expired(update, context):
        return ConversationHandler.END
    context.user_data["maosh"] = update.message.text.strip()
    
    # FINISH - save to Excel
    uid = update.effective_user.id
    closer_score = user_scores.get(uid, 0)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    row = [
        now,
        "COMPLETED",
        str(uid),
        context.user_data.get("username", ""),
        context.user_data.get("ism_familiya", ""),
        context.user_data.get("yosh", ""),
        context.user_data.get("shahar", ""),
        context.user_data.get("tel", ""),
        context.user_data.get("oxirgi_ish", ""),
        context.user_data.get("sohalar", ""),
        context.user_data.get("tajriba_yil", ""),
        context.user_data.get("vazifalar", ""),
        context.user_data.get("boshlash", ""),
        context.user_data.get("maosh", ""),
        context.user_data.get("interview_date", ""),
        f"{closer_score}/18",
    ]
    ok = excel_append_row(row)

    # Admin ga xabar
    admin_msg = (
        "🆕 YANGI ARIZA\n\n"
        f"⏱ {now}\n"
        f"👤 {context.user_data.get('ism_familiya','-')}\n"
        f"🎂 {context.user_data.get('yosh','-')} yosh\n"
        f"📍 {context.user_data.get('shahar','-')}\n"
        f"📞 {context.user_data.get('tel','-')}\n"
        f"💼 Oxirgi ish: {context.user_data.get('oxirgi_ish','-')}\n"
        f"🎯 Sohalar: {context.user_data.get('sohalar','-')}\n"
        f"📊 Tajriba: {context.user_data.get('tajriba_yil','-')}\n"
        f"📝 Vazifalar: {context.user_data.get('vazifalar','-')}\n"
        f"📅 Boshlash: {context.user_data.get('boshlash','-')}\n"
        f"💰 Maosh: {context.user_data.get('maosh','-')}\n"
        f"🗓 Suhbat: {context.user_data.get('interview_date','-')}\n"
        f"👤 @{context.user_data.get('username','')}\n"
        f"⭐️ Closer: {closer_score}/18\n"
        f"🆔 ID: {uid}\n\n"
        f"📄 Excel: {'✅' if ok else '❌'}"
    )
    try:
        await context.bot.send_message(chat_id=ADMIN_ID, text=admin_msg)
    except Exception as e:
        logger.exception("Admin send error: %s", e)

    await update.message.reply_text(
        "✅ Arizangiz qabul qilindi!\n\n"
        f"⭐️ Closer ball: {closer_score}/18\n\n"
        "Tez orada siz bilan bog'lanishadi."
    )
    return ConversationHandler.END

# ===================== CANCEL =====================
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cancel_timer(context.user_data)
    await update.message.reply_text("Bekor qilindi. /start bosib qaytadan boshlang.")
    return ConversationHandler.END

# ===================== RUN =====================
async def _wait_for_stop_signal():
    loop = asyncio.get_running_loop()
    stop_event = asyncio.Event()

    for sig in (signal.SIGINT, signal.SIGTERM):
        try:
            loop.add_signal_handler(sig, stop_event.set)
        except NotImplementedError:
            pass

    await stop_event.wait()

async def main():
    ensure_excel_file()

    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(start_test, pattern="^start_test$")],
        states={
            SAVOL1: [CallbackQueryHandler(savol1_handler, pattern="^s1_[a-c]$")],
            SAVOL2: [CallbackQueryHandler(savol2_handler, pattern="^s2_[a-d]$")],
            SAVOL3: [CallbackQueryHandler(savol3_handler, pattern="^s3_[a-c]$")],
            SAVOL4: [CallbackQueryHandler(savol4_handler, pattern="^s4_[ab]$")],
            SAVOL5: [CallbackQueryHandler(savol5_handler, pattern="^s5_[ab]$")],
            SAVOL6: [CallbackQueryHandler(savol6_handler, pattern="^int:")],

            ISM_FAMILIYA: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_ism_familiya)],
            YOSH: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_yosh)],
            SHAHAR: [CallbackQueryHandler(get_shahar_callback, pattern="^reg:")],
            TEL: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_tel)],
            OXIRGI_ISH: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_oxirgi_ish)],
            SOHALAR: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_sohalar)],
            TAJRIBA_YIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_tajriba_yil)],
            VAZIFALAR: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_vazifalar)],
            BOSHlash: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_boshlash)],
            MAOSH: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_maosh)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
        per_message=False,
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("excel", send_excel_to_admin))
    app.add_handler(conv)

    await app.initialize()
    await app.start()
    await app.updater.start_polling(drop_pending_updates=True)

    await _wait_for_stop_signal()

    await app.updater.stop()
    await app.stop()
    await app.shutdown()

if __name__ == "__main__":
    asyncio.run(main())

